# pyinstaller -w -F --add-data "youtube1.1.ui;./" 유튜브수집_GUI_ver1.3.py

import os
import sys
import time
from datetime import datetime, timedelta
import traceback
import random

from PyQt5.QtWidgets import *
from PyQt5 import uic
from PyQt5.QtGui import *
from PyQt5.QtCore import Qt, QTimer, QTime, QDate
from PyQt5.QtTest import *

from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from openpyxl.styles import Font, Alignment

from bs4 import BeautifulSoup
from selenium.webdriver.chrome.service import Service
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.alert import Alert
from selenium.common.exceptions import UnexpectedAlertPresentException
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver import ActionChains
from urllib.parse import urlparse, parse_qs
import re
import urllib.parse

from youtube_transcript_api import YouTubeTranscriptApi
from youtube_transcript_api._errors import TranscriptsDisabled, NoTranscriptFound

if getattr(sys, 'frozen', False):
    #test.exe로 실행한 경우,test.exe를 보관한 디렉토리의 full path를 취득
    program_directory = os.path.dirname(os.path.abspath(sys.executable))
else:
    #python test.py로 실행한 경우,test.py를 보관한 디렉토리의 full path를 취득
    program_directory = os.path.dirname(os.path.abspath(__file__))

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

UI_PATH = "youtube1.1.ui"

save_date = datetime.today().strftime("%Y%m%d_%H%M")
favicon_image = program_directory + "\\data\icon-16x16.png"
yj_image = program_directory + "\data\icon.png"
exl_sample = program_directory + "\data\\result_sample.xlsx"
scriptidx_file = program_directory + f"\\data\\scriptidx.ini"
setting_file_1 = program_directory + f"\\data\\setting_1.ini"
setting_file_2 = program_directory + f"\\data\\setting_2.ini"
setting_file_3 = program_directory + f"\\data\\setting_3.ini"

class MainDialog(QDialog):
    def __init__(self):
        QDialog.__init__(self, None)

        self.setWindowFlags(Qt.WindowCloseButtonHint | Qt.WindowMaximizeButtonHint | Qt.WindowMinimizeButtonHint)  # 최소화 버튼
        uic.loadUi(os.path.join(BASE_DIR, UI_PATH), self)

        self.setWindowTitle('유튜브 수집 프로그램 Ver 1.3')
        self.setWindowIcon(QIcon(favicon_image))
        pm = QPixmap(yj_image)
        pm = pm.scaledToWidth(220)
        self.yj_image_btn.setPixmap(pm)

        # 업로드날짜 필터 라디오 버튼 그룹화
        self.ud_button_group = QButtonGroup(self)
        self.ud_button_group.addButton(self.udfilter_btn_1)
        self.ud_button_group.addButton(self.udfilter_btn_2)
        self.ud_button_group.addButton(self.udfilter_btn_3)
        self.ud_button_group.addButton(self.udfilter_btn_4)
        self.ud_button_group.addButton(self.udfilter_btn_5)
        self.ud_button_group.addButton(self.udfilter_btn_6)

        self.udfilter_btn_1.setChecked(True)

        # 구분 필터 라디오 버튼 그룹화
        self.gb_button_group = QButtonGroup(self)
        self.gb_button_group.addButton(self.gbfilter_btn_1)
        self.gb_button_group.addButton(self.gbfilter_btn_2)

        self.gbfilter_btn_1.setChecked(True)

        # 길이 필터 라디오 버튼 그룹화
        self.len_button_group = QButtonGroup(self)
        self.len_button_group.addButton(self.lenfilter_btn_1)
        self.len_button_group.addButton(self.lenfilter_btn_2)
        self.len_button_group.addButton(self.lenfilter_btn_3)
        self.len_button_group.addButton(self.lenfilter_btn_4)

        self.lenfilter_btn_1.setChecked(True)

        # 증가추이 라디오 버튼 그룹화
        self.button_group = QButtonGroup(self)
        self.button_group.addButton(self.grp1_btn)
        self.button_group.addButton(self.grp2_btn)

        # 등록일자 중 시작일자 - 현재 연도의 1월 1일로 설정
        today = QDate.currentDate()
        first_day_of_year = QDate(today.year(), 1, 1)
        self.startdate_btn.setDate(first_day_of_year)
        self.enddate_btn.setDate(QDate.currentDate())

        self.layout = QVBoxLayout(self.scrollArea)

        # Scroll Area
        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)
        self.scroll_widget = QWidget()
        self.scroll_layout = QVBoxLayout(self.scroll_widget)
        self.scroll_area.setWidget(self.scroll_widget)
      
        self.layout.addWidget(self.scroll_area)

        self.rows = []
        for _ in range(100):
            self.add_row()

        self.add_btn.clicked.connect(self.add_row) # 폴더추가 버튼 이벤트
        self.start_btn.clicked.connect(self.main) # 수집시작 버튼 이벤트
        self.exlload_btn.clicked.connect(self.exl_load) # 엑셀 Import 버튼 이벤트
        self.folder_reset_btn.clicked.connect(self.folder_reset) # 폴더명 리셋 버튼 이벤트
        self.vidiq_btn.clicked.connect(self.vidiq_folder) # vidIQ 설치폴더 버튼 이벤트
        self.folder_sel_btn.clicked.connect(self.folder_path) # 저장 기본폴더 버튼 이벤트
        self.setting_save_btn.clicked.connect(self.setting_save) # 설정저장

        if os.path.exists(setting_file_1):

            with open(setting_file_1, 'r') as file:
                lines = file.readlines()

            try:
                if lines[0].strip() == "True" :
                    self.udfilter_btn_1.setChecked(True)
            except :
                pass
            try:
                if lines[1].strip() == "True" :
                    self.udfilter_btn_2.setChecked(True)
            except :
                pass
            try:
                if lines[2].strip() == "True" :
                    self.udfilter_btn_3.setChecked(True)
            except :
                pass
            try:
                if lines[3].strip() == "True" :
                    self.udfilter_btn_4.setChecked(True)
            except :
                pass
            try:
                if lines[4].strip() == "True" :
                    self.udfilter_btn_5.setChecked(True)
            except :
                pass
            try:
                if lines[5].strip() == "True" :
                    self.udfilter_btn_6.setChecked(True)
            except :
                pass

            try:
                if lines[6].strip() == "True" :
                    self.gbfilter_btn_1.setChecked(True)
            except :
                pass
            try:
                if lines[7].strip() == "True" :
                    self.gbfilter_btn_2.setChecked(True)
            except :
                pass

            try:
                if lines[8].strip() == "True" :
                    self.lenfilter_btn_1.setChecked(True)
            except :
                pass
            try:
                if lines[9].strip() == "True" :
                    self.lenfilter_btn_2.setChecked(True)
            except :
                pass
            try:
                if lines[10].strip() == "True" :
                    self.lenfilter_btn_3.setChecked(True)
            except :
                pass
            try:
                if lines[11].strip() == "True" :
                    self.lenfilter_btn_4.setChecked(True)
            except :
                pass

            try:
                self.id_btn.setText(lines[12].strip())
            except:
                self.id_btn.setText('')
            try:
                self.pw_btn.setText(lines[13].strip())
            except:
                self.pw_btn.setText('')

            try:
                self.vidiq_path_btn.setText(lines[14].strip())
            except:
                self.vidiq_path_btn.setText('')
            try:
                self.folder_path_btn.setText(lines[15].strip())
            except:
                self.folder_path_btn.setText('')

            try:
                self.limitcnt_btn.setText(lines[16].strip())
            except:
                self.limitcnt_btn.setText('')
            try:
                self.viewcnt_btn.setText(lines[17].strip())
            except:
                self.viewcnt_btn.setText('')

            try:
                selected_date_1 = QDate.fromString(lines[18].strip(), "yyyy-MM-dd")
                self.startdate_btn.setDate(selected_date_1)
            except:
                pass
            try:
                selected_date_2 = QDate.fromString(lines[19].strip(), "yyyy-MM-dd")
                self.enddate_btn.setDate(selected_date_2)
            except:
                pass
            try:
                self.len_start_btn.setText(lines[20].strip())
            except:
                self.len_start_btn.setText('')
            try:
                self.len_end_btn.setText(lines[21].strip())
            except:
                self.len_end_btn.setText('')

            try:
                self.member_start_btn.setText(lines[22].strip())
            except:
                self.member_start_btn.setText('')
            try:
                self.member_end_btn.setText(lines[23].strip())
            except:
                self.member_end_btn.setText('')

            try:
                if lines[24].strip() == "True" :
                    self.grp1_btn.setChecked(True)
            except :
                pass
            try:
                if lines[25].strip() == "True" :
                    self.grp2_btn.setChecked(True)
            except :
                pass
            try:
                self.grp_num_btn.setText(lines[26].strip())
            except:
                self.grp_num_btn.setText('')

            try:
                self.delay_start_btn.setText(lines[27].strip())
            except:
                self.delay_start_btn.setText('')
            try:
                self.delay_end_btn.setText(lines[28].strip())
            except:
                self.delay_end_btn.setText('')
            try:
                self.exl_path_btn.setText(lines[29].strip())
                if lines[29].strip() != '' :
                    self.load_excel_data(lines[29].strip())
            except:
                self.exl_path_btn.setText('')

        if os.path.exists(setting_file_2):

            with open(setting_file_2, 'r') as file:
                lines = file.readlines()
            try:
                self.cnname_except_btn.setPlainText(''.join(lines))
            except:
                self.cnname_except_btn.setPlainText('')

        if os.path.exists(setting_file_3):

            with open(setting_file_3, 'r') as file:
                lines = file.readlines()
            try:
                self.title_except_btn.setPlainText(''.join(lines))
            except:
                self.title_except_btn.setPlainText('')

        # X 버튼 클릭 시 종료 이벤트 처리
        self.setAttribute(Qt.WA_DeleteOnClose)
        self.closeEvent = self.on_closing

    # 프로그램 종료
    def on_closing(self, event):
        try :
            print('종료')
            reply = QMessageBox.question(self, '종료', '프로그램을 종료하시겠습니까?',
                                        QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply == QMessageBox.Yes:
                try :
                    self.driver.close()
                except :
                    pass

                event.accept()
            else:
                event.ignore()
        except Exception:
            err = traceback.format_exc()
            self.textEdit_item.append(f"오류가 발생했습니다. {err}")

    def setting_save(self) :

        udfilter_1 = self.udfilter_btn_1.isChecked()
        udfilter_2 = self.udfilter_btn_2.isChecked()
        udfilter_3 = self.udfilter_btn_3.isChecked()
        udfilter_4 = self.udfilter_btn_4.isChecked()
        udfilter_5 = self.udfilter_btn_5.isChecked()
        udfilter_6 = self.udfilter_btn_6.isChecked()

        gbfilter_1 = self.gbfilter_btn_1.isChecked()
        gbfilter_2 = self.gbfilter_btn_2.isChecked()

        lenfilter_1 = self.lenfilter_btn_1.isChecked()
        lenfilter_2 = self.lenfilter_btn_2.isChecked()
        lenfilter_3 = self.lenfilter_btn_3.isChecked()
        lenfilter_4 = self.lenfilter_btn_4.isChecked()

        vidiq_id = self.id_btn.text()
        vidiq_pw = self.pw_btn.text()

        vidiq_path = self.vidiq_path_btn.text()
        folder_path = self.folder_path_btn.text()

        limit_cnt = self.limitcnt_btn.text()
        upper_viewcnt = self.viewcnt_btn.text()

        wish_date_start = self.startdate_btn.date()
        wish_date_end = self.enddate_btn.date()

        wish_date_start = wish_date_start.toString("yyyy-MM-dd") 
        wish_date_end = wish_date_end.toString("yyyy-MM-dd") 

        len_start = self.len_start_btn.text()
        len_end = self.len_end_btn.text()
        wishsubcnt_start = self.member_start_btn.text()
        wishsubcnt_end = self.member_end_btn.text()

        vidiq_search_1 = self.grp1_btn.isChecked()
        vidiq_search_2 = self.grp2_btn.isChecked()
        grp_num = self.grp_num_btn.text()

        rd_time_start = self.delay_start_btn.text()
        rd_time_end = self.delay_end_btn.text()

        exl_path = self.exl_path_btn.text()

        with open(setting_file_1, 'w') as file:
            file.write(str(udfilter_1) + '\n')
            file.write(str(udfilter_2) + '\n')
            file.write(str(udfilter_3) + '\n')
            file.write(str(udfilter_4) + '\n')
            file.write(str(udfilter_5) + '\n')
            file.write(str(udfilter_6) + '\n')

            file.write(str(gbfilter_1) + '\n')
            file.write(str(gbfilter_2) + '\n')

            file.write(str(lenfilter_1) + '\n')
            file.write(str(lenfilter_2) + '\n')
            file.write(str(lenfilter_3) + '\n')
            file.write(str(lenfilter_4) + '\n')

            file.write(str(vidiq_id) + '\n')
            file.write(str(vidiq_pw) + '\n')
            file.write(str(vidiq_path) + '\n')
            file.write(str(folder_path) + '\n')
            file.write(str(limit_cnt) + '\n')
            file.write(str(upper_viewcnt) + '\n')
            file.write(str(wish_date_start) + '\n')
            file.write(str(wish_date_end) + '\n')
            file.write(str(len_start) + '\n')
            file.write(str(len_end) + '\n')
            file.write(str(wishsubcnt_start) + '\n')
            file.write(str(wishsubcnt_end) + '\n')
            file.write(str(vidiq_search_1) + '\n')
            file.write(str(vidiq_search_2) + '\n')
            file.write(str(grp_num) + '\n')
            file.write(str(rd_time_start) + '\n')
            file.write(str(rd_time_end) + '\n')
            file.write(str(exl_path))

        channel_ecp_keywords = self.cnname_except_btn.toPlainText() # 추출 키워드

        with open(setting_file_2, 'w') as file:
            file.write(channel_ecp_keywords) 

        title_ecp_keywords = self.title_except_btn.toPlainText() # 추출 키워드

        with open(setting_file_3, 'w') as file:
            file.write(title_ecp_keywords) 

        self.textEdit_item.appendPlainText(f"현재 설정이 지정되었습니다.")
        QApplication.processEvents()

    # vidIQ 설치폴더 버튼 이벤트
    def vidiq_folder(self) :
        # Directory 를 선택합니다.
        fname = QFileDialog.getExistingDirectory(self, "Select Directory")
        self.vidiq_path_btn.setText(fname) 

        self.textEdit_item.appendPlainText(f"vidIQ 설치폴더가 지정되었습니다.")
        QApplication.processEvents()

    # 저장 기본폴더 버튼 이벤트
    def folder_path(self) :
        # Directory 를 선택합니다.
        fname = QFileDialog.getExistingDirectory(self, "Select Directory")
        self.folder_path_btn.setText(fname) 

        self.textEdit_item.appendPlainText(f"저장 기본폴더가 지정되었습니다.")
        QApplication.processEvents()

    # 폴더명 리셋
    def folder_reset(self):
        self.exl_path_btn.clear()

        for folder_name_edit, keyword_text_edit in self.rows:
            folder_name_edit.clear()
            keyword_text_edit.clear()

    # 엑셀 Import
    def exl_load(self):
        fname = QFileDialog.getOpenFileName(self, "File Load", program_directory, 'Excel File (*.xlsx)')

        if fname[0]:
            file_name = fname[0]
            self.exl_path_btn.setText(file_name)

            self.load_excel_data(file_name)

            self.textEdit_item.appendPlainText(f"{file_name}을 불러왔습니다.")
            QApplication.processEvents()

        else:
            self.textEdit_item.appendPlainText("파일을 다시 선택해주세요.")
            QApplication.processEvents()

    # 엑셀 데이타 GUI 화면 반영
    def load_excel_data(self, file_name):
        wb = load_workbook(file_name)
        sheet = wb.active
        
        row_index = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):  # 첫 번째 행(헤더 제외)
            folder_name_value = str(row[0]).strip() if row[0] else ""  # A열 (폴더명)
            keyword_text_value = str(row[1]).strip() if row[1] else ""  # B열 (키워드)

            if row_index < len(self.rows):
                folder_name_edit, keyword_text_edit = self.rows[row_index] 
                folder_name_edit.setText(folder_name_value)  # QLineEdit에 폴더명 입력
                keyword_text_edit.setText(keyword_text_value)  # QTextEdit에 키워드 입력
            else:
                self.add_row()  # 새 행 추가
                folder_name_edit, keyword_text_edit = self.rows[-1]  # 방금 추가된 마지막 행 가져오기
                folder_name_edit.setText(folder_name_value)  # QLineEdit에 폴더명 입력
                keyword_text_edit.setText(keyword_text_value)  # QTextEdit에 키워드 입력
        
            row_index += 1

    # 폴더추가
    def add_row(self):
        row_frame = QFrame()
        row_layout = QHBoxLayout(row_frame)
        
        folder_name = QLineEdit()
        folder_name.setFixedWidth(120)
        folder_name.setFixedHeight(30)
        keyword_text = QTextEdit()

        row_layout.addWidget(folder_name)
        row_layout.addWidget(keyword_text)
        
        self.scroll_layout.addWidget(row_frame)
        self.rows.append((folder_name, keyword_text))

        self.scroll_widget.setLayout(self.scroll_layout)

    # 수집시작
    def main(self):

        # YouTube URL에서 동영상 ID 추출 함수
        def extract_video_id(youtube_url):
            video_id = None
            short_url_match = re.search(r"youtu\.be/([a-zA-Z0-9_-]{11})", youtube_url)
            if short_url_match:
                video_id = short_url_match.group(1)
            if not video_id:
                long_url_match = re.search(r"v=([a-zA-Z0-9_-]{11})", youtube_url)
                if long_url_match:
                    video_id = long_url_match.group(1)
            if video_id:
                return video_id
            else:
                raise ValueError("Invalid YouTube URL")

        # YouTube 동영상 ID에서 자막을 추출하는 함수
        def get_transcript(video_id):
            try:
                # 지원되는 자막 목록 가져오기
                transcript_list = YouTubeTranscriptApi.list_transcripts(video_id)
                
                # 사용 가능한 언어 확인
                available_languages = [t.language_code for t in transcript_list]

                # 한국어('ko') 자막 우선 시도
                if 'ko' in available_languages:
                    transcript = YouTubeTranscriptApi.get_transcript(video_id, languages=['ko'])
                    return transcript

                # 한국어가 없으면, 가장 첫 번째 지원되는 언어로 가져오기
                elif available_languages:
                    best_language = available_languages[0]  # 첫 번째 지원되는 언어 선택
                    transcript = YouTubeTranscriptApi.get_transcript(video_id, languages=[best_language])
                    return transcript

                else:
                    return {'error': 'No available subtitles for this video.'}

            except TranscriptsDisabled:
                return {'error': 'Subtitles are disabled for this video.'}
            except NoTranscriptFound:
                return {'error': 'No transcript found for the video in any language.'}
            except Exception as e:
                return {'error': str(e)}

        # 추출된 자막을 하나의 텍스트로 병합하는 함수
        def merge_transcript(transcript):
            transcript_text = " ".join([item['text'] for item in transcript])
            return transcript_text

        # 구독자수 숫자변환
        def convert_subscriber_count(subscriber_str):
            multipliers = {'억': 100000000, '만': 10000, '천': 1000}
            subscriber_str = subscriber_str.replace(' ', '')
            
            for suffix, multiplier in multipliers.items():
                if subscriber_str.endswith(suffix):
                    numeric_part = subscriber_str[:-len(suffix)]
                    return round(float(numeric_part) * multiplier)  # 🔹 `round()` 추가
            
            try:
                return round(float(subscriber_str))  # 🔹 `round()` 추가
            except ValueError:
                return 0

        # 영상길이 초 단위 환산
        def convert_to_seconds(time_str):
            parts = list(map(int, time_str.split(":")))  # ":" 기준으로 나누고 정수 변환
            if len(parts) == 3:  # HH:MM:SS 형식
                hours, minutes, seconds = parts
            elif len(parts) == 2:  # MM:SS 형식
                hours = 0
                minutes, seconds = parts
            elif len(parts) == 1:  # SS 형식 (예: "45")
                hours = 0
                minutes = 0
                seconds = parts[0]
            else:
                return None  # 잘못된 형식 처리
            
            return hours * 3600 + minutes * 60 + seconds

        # 조회수 숫자변환
        def convert_views(view_str):
            match = re.match(r"([\d\.]+)([천만억]*)회?", view_str)
            
            if not match:
                return None  # 매칭되지 않으면 None 반환
            
            num, unit = match.groups()
            num = float(num)  # 숫자 부분 변환
            
            # 단위별 변환
            unit_multipliers = {"천": 1_000, "만": 10_000, "억": 100_000_000}
            multiplier = unit_multipliers.get(unit, 1)  # 기본값 1 (단위 없음)
            
            return int(num * multiplier)  # 최종 변환 값

        try :

            vidiq_id = self.id_btn.text()
            vidiq_pw = self.pw_btn.text()

            vidiq_path = self.vidiq_path_btn.text()
            folder_path = self.folder_path_btn.text()

            limit_cnt = self.limitcnt_btn.text()
            upper_viewcnt = self.viewcnt_btn.text()

            wish_date_start = self.startdate_btn.date()
            wish_date_end = self.enddate_btn.date()

            wish_date_start = wish_date_start.toString("yyyy-MM-dd") 
            wish_date_end = wish_date_end.toString("yyyy-MM-dd") 

            len_start = self.len_start_btn.text()
            len_end = self.len_end_btn.text()
            wishsubcnt_start = self.member_start_btn.text()
            wishsubcnt_end = self.member_end_btn.text()

            vidiq_search_1 = self.grp1_btn.isChecked()
            vidiq_search_2 = self.grp2_btn.isChecked()
            grp_num = self.grp_num_btn.text()

            rd_time_start = self.delay_start_btn.text()
            rd_time_end = self.delay_end_btn.text()

            exl_path = self.exl_path_btn.text()

            channel_ecp_keywords = self.cnname_except_btn.toPlainText()
            title_ecp_keywords = self.title_except_btn.toPlainText() 

            udfilter_check_1 = self.udfilter_btn_1.isChecked()
            udfilter_check_2 = self.udfilter_btn_2.isChecked()
            udfilter_check_3 = self.udfilter_btn_3.isChecked()
            udfilter_check_4 = self.udfilter_btn_4.isChecked()
            udfilter_check_5 = self.udfilter_btn_5.isChecked()
            udfilter_check_6 = self.udfilter_btn_6.isChecked()

            gbfilter_check_1 = self.gbfilter_btn_1.isChecked()
            gbfilter_check_2 = self.gbfilter_btn_2.isChecked()

            lenfilter_check_1 = self.lenfilter_btn_1.isChecked()
            lenfilter_check_2 = self.lenfilter_btn_2.isChecked()
            lenfilter_check_3 = self.lenfilter_btn_3.isChecked()
            lenfilter_check_4 = self.lenfilter_btn_4.isChecked()

            if udfilter_check_2 :
                ud_text = '지난 1시간'
            if udfilter_check_3 :
                ud_text = '오늘'
            if udfilter_check_4 :
                ud_text = '이번 주'
            if udfilter_check_5 :
                ud_text = '이번 달'
            if udfilter_check_6 :
                ud_text = '올해'

            if gbfilter_check_2 :
                gb_text = '동영상'

            if lenfilter_check_2 :
                len_text = '4분 미만'
            if lenfilter_check_3 :
                len_text = '4~20분'
            if lenfilter_check_4 :
                len_text = '20분 초과'


            # 유효성 검사

            if vidiq_id == '' :
                QMessageBox.information(self, "설정오류", "vidIQ ID를 입력하세요.")
                QApplication.processEvents()
                return 0
            if vidiq_pw == '' :
                QMessageBox.information(self, "설정오류", "vidIQ PW를 입력하세요.")
                QApplication.processEvents()
                return 0
            if vidiq_path == '' :
                QMessageBox.information(self, "설정오류", "vidIQ 설치폴더를 입력하세요.")
                QApplication.processEvents()
                return 0
            if folder_path == '' :
                QMessageBox.information(self, "설정오류", "저장 기본폴더를 입력하세요.")
                QApplication.processEvents()
                return 0

            try :
                limit_cnt = int(limit_cnt)
            except :
                QMessageBox.information(self, "설정오류", "키워드당 수집개수를 숫자로 입력하세요.")
                QApplication.processEvents()
                return 0

            try :
                upper_viewcnt = int(upper_viewcnt)
            except :
                QMessageBox.information(self, "설정오류", "조회수(이상)를 숫자로 입력하세요.")
                QApplication.processEvents()
                return 0    

            if wish_date_start > wish_date_end :
                QMessageBox.information(self, "설정오류", "시작일자를 종료일자보다 빠른 일자로 설정하세요.")
                QApplication.processEvents()
                return 0 

            try :
                len_start = int(len_start)
            except :
                QMessageBox.information(self, "설정오류", "영상길이(분) 시작 값을 숫자로 입력하세요.")
                QApplication.processEvents()
                return 0 
            try :
                len_end = int(len_end)
            except :
                QMessageBox.information(self, "설정오류", "영상길이(분) 종료 값을 숫자로 입력하세요.")
                QApplication.processEvents()
                return 0 
            if len_start > len_end :  
                QMessageBox.information(self, "설정오류", "영상길이(분) 시작값을 종료값 보다 작게 입력하세요.")
                QApplication.processEvents()
                return 0

            try :
                wishsubcnt_start = int(wishsubcnt_start)
            except :
                QMessageBox.information(self, "설정오류", "채널구독자수(명) 시작 값을 숫자로 입력하세요.")
                QApplication.processEvents()
                return 0 
            try :
                wishsubcnt_end = int(wishsubcnt_end)
            except :
                QMessageBox.information(self, "설정오류", "채널구독자수(명) 종료 값을 숫자로 입력하세요.")
                QApplication.processEvents()
                return 0 
            if wishsubcnt_start > wishsubcnt_end :  
                QMessageBox.information(self, "설정오류", "채널구독자수(명) 시작값을 종료값 보다 작게 입력하세요.")
                QApplication.processEvents()
                return 0

            if vidiq_search_1 == False and vidiq_search_2 == False :
                QMessageBox.information(self, "설정오류", "조회수 증가추세를 설정하세요.")
                QApplication.processEvents()
                return 0 
            if vidiq_search_1 and grp_num == '' :
                QMessageBox.information(self, "설정오류", "증가추세만 가져오기 선택시 몇개를 갖고올지 설정하세요.")
                QApplication.processEvents()
                return 0
            if vidiq_search_1 :
                try :
                    grp_num = int(grp_num)
                except :
                    QMessageBox.information(self, "설정오류", "증가추세만 가져오기의 비교대상 지표값 갯수를 숫자로 입력하세요.")
                    QApplication.processEvents()
                    return 0 

                if grp_num < 3 :
                    QMessageBox.information(self, "설정오류", "증가추세만 가져오기의 비교대상 지표값 갯수는 3 이상부터 입력 가능합니다.")
                    QApplication.processEvents()
                    return 0 

            try :
                rd_time_start = int(rd_time_start)
            except :
                QMessageBox.information(self, "설정오류", "딜레이(초) 시작 값을 숫자로 입력하세요.")
                QApplication.processEvents()
                return 0 
            try :
                rd_time_end = int(rd_time_end)
            except :
                QMessageBox.information(self, "설정오류", "딜레이(초) 종료 값을 숫자로 입력하세요.")
                QApplication.processEvents()
                return 0
            if rd_time_start > rd_time_end :  
                QMessageBox.information(self, "설정오류", "딜레이(초) 시작값을 종료값 보다 작게 입력하세요.")
                QApplication.processEvents()
                return 0

            subfolders = [os.path.join(vidiq_path, d) for d in os.listdir(vidiq_path) if os.path.isdir(os.path.join(vidiq_path, d))]

            # 가장 최근에 생성된 폴더 찾기
            if subfolders:
                vidiq_latest_folder = max(subfolders, key=os.path.getctime)
            else:
                QMessageBox.information(self, "설정오류", "vidIQ 확장프로그램에 최근 설치된 폴더가 존재하지 않습니다.")
                QApplication.processEvents()
                return 0

            folder_lines = []
            
            for folder_name, keyword_text in self.rows:
                folder_value = folder_name.text().strip()
                keyword_lines = [line.strip() for line in keyword_text.toPlainText().split("\n") if line.strip()]  # 공란 제거

                if folder_value == '' :
                    break
                
                if keyword_lines:  # keyword_text가 비어 있지 않은 경우만 추가
                    folder_lines.append([folder_value] + keyword_lines)

            if len(folder_lines) == 0 :
                QMessageBox.information(self, "설정오류", "수집진행할 폴더명/키워드명이 존재하지 않습니다.(키워드명 입력시 폴더명 필수값)")
                QApplication.processEvents()
                return 0

            now_time = datetime.today().strftime("%Y-%m-%d_%H:%M")
            self.textEdit_item.appendPlainText(f'\n▶▶▶ 유튜브 정보수집을 시작합니다.({now_time})\n')
            QApplication.processEvents()

            # ChromeOptions 설정
            options = webdriver.ChromeOptions()
            options.add_argument("--disable-blink-features=AutomationControlled")
            options.add_argument('Accept-Language=ko-KR,ko;q=0.8,en-US;q=0.5,en;q=0.3')
            options.add_argument("--no-sandbox")
            options.add_argument("--disable-gpu")
            options.add_argument("--disable-page-load-metrics")
            options.add_argument("--disable-devtools-experiments")
            options.add_argument("--disable-geolocation")
            options.add_experimental_option("useAutomationExtension", False)
            options.add_experimental_option("excludeSwitches", ["enable-automation"])
            options.add_argument("--disable-dev-shm-usage")
            options.add_argument(f"--force-device-scale-factor=0.7")
            options.add_experimental_option("prefs", {
                "credentials_enable_service": False, 
                "profile.password_manager_enabled": False  
            })

            if vidiq_search_1 :
                options.add_argument(f"--load-extension={vidiq_latest_folder}")

            driver = webdriver.Chrome(options=options)


            if vidiq_search_1 :
                driver.get("https://app.vidiq.com/auth/login")
                time.sleep(3)

                handles = driver.window_handles
                driver.switch_to.window(handles[0])
                time.sleep(1)


                id_btn = driver.find_element(by=By.CSS_SELECTOR, value= f"#email") # 검색어 Input
                id_btn.send_keys(vidiq_id)
                time.sleep(0.2)
                pw_btn = driver.find_element(by=By.CSS_SELECTOR, value= f"#password") # 검색어 Input
                pw_btn.send_keys(vidiq_pw)
                time.sleep(0.2)
                driver.find_element(by=By.CSS_SELECTOR, value= f".chakra-button.css-1jtildb").click()
                time.sleep(1)

                QMessageBox.information(self, "로그인 완료여부", "정상적으로 vidIQ에 로그인이 되었다면 [OK] 버튼을 클릭하세요.(자동로그인 실패시 수동로그인 필요)")
                QApplication.processEvents()

            if len(driver.window_handles) == 1 :
                driver.execute_script("window.open('');")
                time.sleep(1)

            # 수집일자 폴더 세팅
            current_time = datetime.now().strftime("%y%m%d")
            date_folder = folder_path + f'\\{current_time}'

            if not os.path.exists(date_folder):
                os.makedirs(date_folder)

            # 사용자 설정값 세팅
            if vidiq_search_1 :
                vidiq_search = True # 그래프 추이 확인여부
            else :
                vidiq_search = False

            wish_date_start = datetime.strptime(wish_date_start, "%Y-%m-%d").date()
            wish_date_end = datetime.strptime(wish_date_end, "%Y-%m-%d").date()

            wish_len_start = len_start * 60
            wish_len_end = len_end * 60

            channel_ecp_keywords = [line.strip() for line in channel_ecp_keywords.split("\n") if line.strip()]
            title_ecp_keywords = [line.strip() for line in title_ecp_keywords.split("\n") if line.strip()]

            # 엑셀세팅
            save_date = datetime.today().strftime("%Y%m%d_%H%M%S")
            exlsave_name = date_folder + f"\\{save_date}.xlsx"

            wb = load_workbook(exl_sample)
            sheet = wb.active
            last_row = sheet.max_row

            search_num = 1

            # 반복문으로 수집시작
            for folder_line in folder_lines :

                folder_name = folder_line[0]
                self.searchfolder_btn.setText(folder_name)
                QApplication.processEvents()

                user_folder = date_folder + f'\\{folder_name}'

                if not os.path.exists(user_folder):
                    os.makedirs(user_folder)

                for keyword in folder_line[1:] :

                    kwd_search_num = 0

                    self.searchkwd_btn.setText(keyword)
                    QApplication.processEvents()

                    self.textEdit_item.appendPlainText(f'\n▶▶ {keyword}(폴더명 : {folder_name}) 확인 중\n')
                    QApplication.processEvents()

                    driver.switch_to.window(driver.window_handles[0])
                    time.sleep(1)

                    query = urllib.parse.quote(keyword) 

                    link = f'https://www.youtube.com/results?search_query={query}'
                    driver.get(link)
                    time.sleep(5)

                    view_tabs = driver.find_elements(By.CSS_SELECTOR, "#chips > yt-chip-cloud-chip-renderer")

                    for view_tab in view_tabs : 
                        if view_tab.text == '동영상' :
                            view_tab.click()
                            time.sleep(3)

                    # 유튜브 검색필터 추가
                    for i in range(3) :
                        if i == 0 and udfilter_check_1 == True :
                            continue
                        if i == 1 and udfilter_check_1 == False :
                            continue
                        if i == 1 and gbfilter_check_1 == True :
                            continue
                        if i == 2 and lenfilter_check_1 == True :
                            continue

                        if i == 0 :
                            search_txt = ud_text
                        if i == 1 :
                            search_txt = gb_text
                        if i == 2 :
                            search_txt = len_text

                        filter_c_tag = driver.find_element(By.CSS_SELECTOR, ".yt-spec-button-shape-next.yt-spec-button-shape-next--text.yt-spec-button-shape-next--mono.yt-spec-button-shape-next--size-m.yt-spec-button-shape-next--icon-trailing.yt-spec-button-shape-next--enable-backdrop-filter-experiment")
                        filter_c_tag.click()
                        time.sleep(2)

                        filter_tags = driver.find_elements(By.CSS_SELECTOR, "ytd-search-filter-renderer")

                        if i == 0 :
                            search_txt = ud_text
                            rst_filter_tags = filter_tags[:5]
                        if i == 1 :
                            search_txt = gb_text
                            rst_filter_tags = filter_tags[5:9]
                        if i == 2 :
                            search_txt = len_text
                            rst_filter_tags = filter_tags[9:12]

                        for rst_filter_tag in rst_filter_tags :
                            current_tag = rst_filter_tag.find_element(By.CSS_SELECTOR, "#label")
                            current_text = current_tag.text.strip()

                            if current_text == search_txt :
                                current_tag.click()
                                time.sleep(3)
                                break

                    except_link = []

                    while 1 :

                        driver.switch_to.window(driver.window_handles[0])
                        time.sleep(1)

                        # 현재 페이지 높이 저장
                        last_height = driver.execute_script("return document.documentElement.scrollHeight")

                        soup = BeautifulSoup(driver.page_source, "html.parser")
                        videos = soup.select("ytd-video-renderer")

                        video_links = []

                        for idx, video in enumerate(videos,start=1) :

                            descriptions = video.select_one("#metadata-line").text.split("\n")

                            # 제목
                            title_tag = video.select_one("#video-title")
                            title = title_tag.text.strip()

                            # 링크
                            links = title_tag.get("href")
                            links = links.split('&')
                            link = 'https://www.youtube.com' + links[0]

                            if link in except_link :
                                continue

                            except_link.append(link)

                            # 제목 금칙어 해당여부
                            title_pass = None
                            for title_ecp_keyword in title_ecp_keywords :
                                if title_ecp_keyword in title :
                                    title_pass = True
                                    break
                            if title_pass == True :
                                print(f'▷ {title} : 영상제목 금지어 포함되어 수집제외({title_ecp_keyword})[1차 필터링]')
                                self.textEdit_item.appendPlainText(f'▷ {title} : 영상제목 금지어 포함되어 수집제외({title_ecp_keyword})[1차 필터링]')
                                QApplication.processEvents()
                                continue            

                            # 채널명
                            channel_tags = video.select("a.yt-simple-endpoint.style-scope.yt-formatted-string")
                            channel = channel_tags[1].text.strip()

                            # 채널명 금칙어 해당여부
                            cnname_pass = None
                            for channel_ecp_keyword in channel_ecp_keywords  :
                                if channel_ecp_keyword in channel :
                                    cnname_pass = True
                                    break
                            if cnname_pass == True :
                                print(f'▷ {title} : 채널명 금지어 포함되어 수집제외({channel_ecp_keyword})[1차 필터링]')
                                self.textEdit_item.appendPlainText(f'▷ {title} : 채널명 금지어 포함되어 수집제외({channel_ecp_keyword})[1차 필터링]')
                                QApplication.processEvents()
                                continue 

                            # 영상길이
                            duration_tag = video.select_one("ytd-thumbnail-overlay-time-status-renderer span")
                            duration = duration_tag.text.strip() if duration_tag else "길이 정보 없음"

                            if duration == "길이 정보 없음" or duration == "" or duration == "예정" :
                                continue
                            try:
                                sec_lenth = convert_to_seconds(duration)
                            except :
                                continue

                            if not (wish_len_start <= sec_lenth <= wish_len_end) :
                                print(f'▷ {title} : 영상길이 미충족 수집제외({sec_lenth}초)[1차 필터링]')
                                self.textEdit_item.appendPlainText(f'▷ {title} : 영상길이 미충족 수집제외({sec_lenth}초)[1차 필터링]')
                                QApplication.processEvents()
                                continue

                            # 스트리밍 여부
                            upload_date = descriptions[4]
                            if '스트리밍' in upload_date :
                                print(f'▷ {title} : 스트리밍 수집제외[1차 필터링]')
                                self.textEdit_item.appendPlainText(f'▷ {title} : 스트리밍 수집제외[1차 필터링]')
                                QApplication.processEvents()
                                continue

                            # 조회수
                            if ' 없음' not in descriptions[3] :
                                view_cnts = descriptions[3].replace('조회수 ','')
                                view_cnt = convert_views(view_cnts)
                            else :
                                view_cnt = 0

                            if view_cnt < upper_viewcnt : # 조회수 기준 미달시 패스
                                print(f'▷ {title} : 조회수 미충족 수집제외({view_cnt}회)[1차 필터링]')
                                self.textEdit_item.appendPlainText(f'▷ {title} : 조회수 미충족 수집제외({view_cnt}회)[1차 필터링]')
                                QApplication.processEvents()
                                continue

                            video_links.append([title, link, channel])
                            print(f'□ {title} : 1차 링크 수집완료')

                        if len(video_links) > 0 :

                            driver.switch_to.window(driver.window_handles[1])
                            time.sleep(1)

                            for video_link in video_links :

                                try :

                                    ytb_title = video_link[0]
                                    ytb_link = video_link[1]
                                    ytb_channel = video_link[2]

                                    driver.get(ytb_link)
                                    time.sleep(5)

                                    if vidiq_search : 

                                        vid_iq_true = False

                                        vidiq_wait_num =1

                                        while 1:

                                            try :
                                                req = driver.page_source 
                                                soup = BeautifulSoup(req, 'html.parser')


                                                svg_tags = soup.find_all('svg', {'class': 'recharts-surface', 'overflow': 'visible'})
                                                d_attr = svg_tags[-1].select_one('g.recharts-layer.recharts-line path').get('d')

                                                # 정규식으로 숫자 추출
                                                coordinates = re.findall(r"[-+]?\d*\.?\d+", d_attr)

                                                # (x, y) 형태로 변환
                                                points = [(float(coordinates[i]), float(coordinates[i + 1])) for i in range(0, len(coordinates), 2)]
                                                points = points[1:-1]

                                                break
                                            except :
                                                time.sleep(0.5)
                                                try :
                                                    nodata_tag = soup.select_one('.vidiq-c-fvFDqp.vidiq-c-fvFDqp-hyvuql-weight-bold.vidiq-c-fvFDqp-koygTM-size-md.vidiq-c-fvFDqp-ipJykX-css').text.strip()
                                                    if 'No views' in nodata_tag :
                                                        points = []
                                                        break
                                                except :
                                                    pass

                                                vidiq_wait_num += 1

                                                if vidiq_wait_num == 30 :
                                                    points = []
                                                    break
                                                
                                        if len(points) >= grp_num :
                                            try :
                                                # y 좌표값만 추출
                                                y_values = [y for _, y in points]
                                                y_values = y_values[-grp_num:]
                                                differences = [round(y_values[i] - y_values[i + 1], 3) for i in range(len(y_values) - 1)]

                                                # 비교 결과를 저장할 리스트
                                                comparison_results = [differences[i] >= differences[i - 1] for i in range(1, len(differences))]
                                                vid_iq_true = all(comparison_results)
                                            except :
                                                vid_iq_true = False
                                        if len(points) < grp_num :
                                            y_values = ['좌표수 부족']
                                            differences = ['좌표수 부족']
                                            comparison_results = ['좌표수 부족']
                                            vid_iq_true = False

                                        if vid_iq_true == False :

                                            self.textEdit_item.appendPlainText(f'▷ {ytb_title} : vidIQ 지수조건 미충족 수집제외({view_cnt}회)[2차 필터링]')
                                            self.textEdit_item.appendPlainText(f'   - Y좌표 : {y_values} / 좌표별 증감값 : {differences} / 좌표간 증감여부 : {comparison_results}')
                                            QApplication.processEvents()
                                            continue

                                    scroll_num = 1

                                    while True:
                                        try:
                                            more_button = driver.find_element(by=By.CSS_SELECTOR, value= f"ytd-text-inline-expander#description-inline-expander >  tp-yt-paper-button#expand") # 검색어 Input
                                            more_button.click()
                                            time.sleep(2)
                                            break
                                        except:
                                            driver.execute_script("window.scrollBy(0, 50);")
                                            time.sleep(0.5)  # 로딩 대기

                                            scroll_num += 1

                                            if scroll_num == 20 :
                                                break

                                    if scroll_num == 20 :
                                        continue

                                    soup = BeautifulSoup(driver.page_source, "html.parser")

                                    # 업로드 날짜
                                    info_tag = soup.select_one('yt-formatted-string#info')
                                    upload_date = None

                                    if info_tag:
                                        # 모든 span 태그 순회
                                        for span in info_tag.find_all("span"):
                                            text = span.get_text(strip=True)
                                            
                                            if re.match(r"^\d{4}\. ?\d{1,2}\. ?\d{1,2}\.?$", text):
                                                # 끝 마침표 제거 후 datetime 변환
                                                clean_date = text.rstrip(".")
                                                try:
                                                    upload_date_str = datetime.strptime(clean_date, "%Y. %m. %d")
                                                    upload_date = upload_date_str.date()
                                                    break  # 찾았으면 반복 종료
                                                except ValueError:
                                                    continue

                                    if not (wish_date_start <= upload_date <= wish_date_end) :
                                        print(f'▷ {ytb_title} : 등록일자 기준 미충족 수집제외[2차 필터링]')
                                        self.textEdit_item.appendPlainText(f'▷ {ytb_title} : 등록일자 기준 미충족 수집제외({upload_date})[2차 필터링]')
                                        QApplication.processEvents()
                                        continue

                                    # 조회수                                   
                                    info_elems = soup.select('#info')  # 여러 개 선택
                                    for elem in info_elems:
                                        view_count_text = elem.get_text(strip=True)
                                        if '조회수' in text:
                                            break

                                    match = re.search(r'조회수\s*([\d,]+)\s*회', view_count_text)
                                    if match:
                                        view_count = int(match.group(1).replace(',', ''))
                                        print('조회수 : ', view_count)

                                    if view_count < upper_viewcnt :
                                        print(f'▷ {ytb_title} : 조회수 미충족 수집제외({view_count}회)[2차 필터링]')
                                        self.textEdit_item.appendPlainText(f'▷ {ytb_title} : 조회수 미충족 수집제외({view_count}회)[2차 필터링]')
                                        QApplication.processEvents()
                                        continue

                                    # 구독자수
                                    sub_count_txt = soup.select_one('#owner-sub-count').text.strip()
                                    sub_count_txt = sub_count_txt.replace('구독자 ','').replace('명','')
                                    subscribe_cnt = convert_subscriber_count(sub_count_txt)

                                    if not (wishsubcnt_start <= subscribe_cnt <= wishsubcnt_end) :
                                        print(f'▷ {ytb_title} : 구독자수 기준 미충족 수집제외[2차 필터링]')
                                        self.textEdit_item.appendPlainText(f'▷ {ytb_title} : 구독자수 미충족 수집제외({subscribe_cnt}명)[2차 필터링]')
                                        QApplication.processEvents()
                                        continue

                                    # 자막수집 / 자막인덱스번호 업데이트
                                    try :
                                        video_id = extract_video_id(ytb_link)
                                        transcript = get_transcript(video_id)
                                        transcript_text = merge_transcript(transcript)
                                    except :
                                        transcript_text = ''

                                    script_index_str = ''

                                    if transcript_text != '' :

                                        with open(scriptidx_file, 'r') as file:
                                            lines = file.readlines()

                                        # 스크립트 인덱스 파싱 (파일이 비어 있으면 기본값 설정)
                                        script_index = int(lines[0].strip()) if lines else 0

                                        # 인덱스 증가 및 8자리 문자열 변환
                                        script_index += 1
                                        script_index_str = f"{script_index:08d}"

                                        # 변경된 값 다시 저장 (8자리 문자열로 저장)
                                        with open(scriptidx_file, 'w') as file:
                                            file.write(script_index_str)

                                        script_filename = user_folder + f"\\{script_index_str}.txt"

                                        with open(script_filename, 'w', encoding='utf-8') as file:
                                            file.write(transcript_text)

                                    # 엑셀저장
                                    current_date = datetime.now().date()
                                    date_style = NamedStyle(name="datetime", number_format="YYYY-MM-DD")
                                    if "datetime" not in sheet.parent.named_styles:
                                        sheet.parent.add_named_style(date_style)

                                    sheet[f'A{search_num + last_row}'] = folder_name # 폴더명
                                    sheet[f'B{search_num + last_row}'] = keyword # 키워드
                                    sheet[f'C{search_num + last_row}'] = str(script_index_str) # 영상수집번호
                                    sheet[f'D{search_num + last_row}'] = current_date # 수집일자
                                    sheet[f'E{search_num + last_row}'] = view_count # 조회수

                                    exl_link = ytb_link.replace('?v=','/')

                                    sheet[f'F{search_num + last_row}'] = exl_link # 링크
                                    try :
                                        sheet[f'G{search_num + last_row}'] = ytb_channel # 채널명
                                    except :
                                        pass
                                    sheet[f'H{search_num + last_row}'] = int(subscribe_cnt) # 구독자수
                                    try :
                                        sheet[f'I{search_num + last_row}'] = ytb_title # 제목
                                    except :
                                        pass
                                    sheet[f'J{search_num + last_row}'] = upload_date # 업로드 날짜

                                    sheet[f'D{search_num + last_row}'].style = "datetime"
                                    sheet[f'F{search_num + last_row}'].hyperlink = sheet[f'F{search_num + last_row}'].value
                                    sheet[f'F{search_num + last_row}'].style = "Hyperlink"
                                    sheet[f'F{search_num + last_row}'].alignment = Alignment(horizontal="left", vertical="center")
                                    sheet[f'E{search_num + last_row}'].number_format = '#,##0'
                                    sheet[f'H{search_num + last_row}'].number_format = '#,##0'
                                    sheet[f'J{search_num + last_row}'].style = "datetime"

                                    wb.save(exlsave_name)

                                    search_num += 1
                                    kwd_search_num +=1

                                    print(f'▶ {kwd_search_num}/{limit_cnt} - {ytb_title} : 수집완료')
                                    self.textEdit_item.appendPlainText(f'▶ {kwd_search_num}/{limit_cnt} - {ytb_title} : 수집완료')
                                    QApplication.processEvents()

                                    if kwd_search_num == limit_cnt :
                                        break

                                    time.sleep(random.uniform(rd_time_start, rd_time_end))

                                except Exception:
                                    err = traceback.format_exc()
                                    print(err)
                                    pass 


                        driver.switch_to.window(driver.window_handles[0])
                        time.sleep(1)

                        if kwd_search_num == limit_cnt :
                            break

                        # 페이지 끝까지 스크롤
                        driver.find_element(By.TAG_NAME, "body").send_keys(Keys.END)

                        # 로딩 대기
                        time.sleep(5)
                        

                        # 새로운 페이지 높이 가져오기
                        new_height = driver.execute_script("return document.documentElement.scrollHeight")

                        # 더 이상 스크롤할 곳이 없으면 종료
                        if new_height == last_height:
                            break

                    time.sleep(1)

            now_time = datetime.today().strftime("%Y-%m-%d_%H:%M")
            self.textEdit_item.appendPlainText(f'\n▶▶▶ 유튜브 정보수집이 완료되었습니다.({now_time})\n')
            QApplication.processEvents()

            QMessageBox.information(self, "수집완료", "유튜브 정보수집이 완료되었습니다.")
            QApplication.processEvents()

        except Exception:
            err = traceback.format_exc()
            self.textEdit_item.appendPlainText(err) 
            now_time = datetime.today().strftime("%Y-%m-%d_%H:%M")
            self.textEdit_item.appendPlainText(f'\n▶▶▶ 오류가 발생되어 정보수집이 중지되었습니다.({now_time})\n')
            QApplication.processEvents()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainDialog()
    window.show()
    sys.exit(app.exec_())